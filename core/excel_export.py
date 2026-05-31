"""
Generador de Excel propio de calificaciones.
Formato: encabezado con dimensiones + tabla de notas.
Las columnas ProT / ConT / ActT se dejan VACÍAS intencionalmente
para poder copiar la fila directo a la plataforma del colegio.
"""
import io
from datetime import date
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# Columnas fijas de la plantilla del colegio (en orden)
# grupo: "proc" = PROCEDIMENTAL, "conc" = CONCEPTUAL, "act" = ACTITUDINAL
# skip=True significa que la celda se deja vacía (ProT, ConT, ActT)
COLUMNAS_FIJAS = [
    # abrev,  grupo,    skip
    ("Taller", "proc",  False),
    ("Tareas", "proc",  False),
    ("Plat",   "proc",  False),
    ("Cuad",   "proc",  False),
    ("ProT",   "proc",  True),   # ← vacío
    ("Exp",    "conc",  False),
    ("Quiz",   "conc",  False),
    ("EvaO",   "conc",  False),
    ("Part",   "conc",  False),
    ("ConT",   "conc",  True),   # ← vacío
    ("Disc",   "act",   False),
    ("Auto",   "act",   False),
    ("ActT",   "act",   True),   # ← vacío
]

# Mapa: nombre de categoría en la app → índice en COLUMNAS_FIJAS (0-based)
NOMBRE_A_IDX = {
    "Tall":  0,
    "Tar.":  1,
    "Plat":  2,
    "Cuad":  3,
    "Exp.":  5,
    "Quiz":  6,
    "EvaO":  7,
    "Part":  8,
    "Disc":  10,
    "Auto":  11,
}

# Colores
AZUL_HEADER  = "FF17375E"   # azul oscuro (encabezado principal)
AZUL_MEDIO   = "FF4F81BD"   # azul medio (sub-encabezado)
AZUL_CLARO   = "FFB8CCE4"   # azul claro (% y col-headers)
SKIP_COLOR   = "FFD9D9D9"   # gris (ProT / ConT / ActT)
BLANCO       = "FFFFFFFF"
NEGRO        = "FF000000"


def _font(bold=False, color=NEGRO, size=10, name="Calibri"):
    return Font(name=name, bold=bold, size=size, color=color)

def _fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def _border():
    thin = Side(style="thin", color="FF000000")
    return Border(left=thin, right=thin, top=thin, bottom=thin)

def _center():
    return Alignment(horizontal="center", vertical="center", wrap_text=True)

def _left():
    return Alignment(horizontal="left", vertical="center")


def generar_excel_propio(materia, periodo, estudiantes, categorias, notas_dict):
    """
    Genera Excel propio de notas.

    Args:
        materia:      objeto Materia
        periodo:      int 1-4
        estudiantes:  iterable de Estudiante activos (ordenados por nombre)
        categorias:   iterable de CategoriaNota del periodo
        notas_dict:   {(est_id, cat_id): valor_int_o_None}

    Returns:
        BytesIO con el .xlsx
    """
    wb = Workbook()
    ws = wb.active
    ws.title = f"Notas P{periodo}"

    hoy = date.today()
    brd = _border()

    # ── Mapear categorías de la app a columnas de la plantilla ───
    # cat_col_map: {cat.id: idx_en_COLUMNAS_FIJAS}
    cat_col_map = {}
    for cat in categorias:
        nombre = cat.nombre.strip()
        if nombre in NOMBRE_A_IDX:
            cat_col_map[cat.id] = NOMBRE_A_IDX[nombre]
        else:
            for k, v in NOMBRE_A_IDX.items():
                if k.lower() in nombre.lower():
                    cat_col_map[cat.id] = v
                    break

    # Columnas de datos: A=Código, B=Estudiante, C=No., luego 13 columnas notas
    # → columna D = COLUMNAS_FIJAS[0], E=[1], ... P=[12]
    DATA_START = 4   # columna D (1-indexed)
    NUM_NOTE_COLS = len(COLUMNAS_FIJAS)  # 13

    # ── FILA 1: Encabezado institución ──────────────────────────
    ws.row_dimensions[1].height = 18
    ws["A1"] = "Jornada:"
    ws["A1"].font = _font(bold=True, size=10)
    ws["B1"] = "U"
    ws["B1"].font = _font(bold=True, color="FF0070C0", size=10)
    ws["C1"] = ""
    ws["D1"] = "Grado:"
    ws["D1"].font = _font(bold=True, size=10)
    ws["E1"] = materia.grado.nombre
    ws["E1"].font = _font(bold=True, color="FF0070C0", size=10)
    ws["F1"] = ""

    # Dimensiones PROCEDIMENTAL (cols D-H = índices 0-4)
    proc_start = DATA_START
    proc_end   = DATA_START + 4
    conc_start = DATA_START + 5
    conc_end   = DATA_START + 9
    act_start  = DATA_START + 10
    act_end    = DATA_START + 12

    ws.merge_cells(start_row=1, start_column=proc_start, end_row=1, end_column=proc_end)
    ws.cell(1, proc_start).value = "PROCEDIMENTAL"
    ws.cell(1, proc_start).font  = _font(bold=True, color=BLANCO, size=10)
    ws.cell(1, proc_start).fill  = _fill(AZUL_HEADER)
    ws.cell(1, proc_start).alignment = _center()

    ws.merge_cells(start_row=1, start_column=conc_start, end_row=1, end_column=conc_end)
    ws.cell(1, conc_start).value = "30%    CONCEPTUAL"
    ws.cell(1, conc_start).font  = _font(bold=True, color=BLANCO, size=10)
    ws.cell(1, conc_start).fill  = _fill(AZUL_HEADER)
    ws.cell(1, conc_start).alignment = _center()

    ws.merge_cells(start_row=1, start_column=act_start, end_row=1, end_column=act_end)
    ws.cell(1, act_start).value = "60%    ACTITUDINAL"
    ws.cell(1, act_start).font  = _font(bold=True, color=BLANCO, size=10)
    ws.cell(1, act_start).fill  = _fill(AZUL_HEADER)
    ws.cell(1, act_start).alignment = _center()

    # ── FILA 2: Escala + porcentajes por columna ─────────────────
    ws.row_dimensions[2].height = 22
    ws["A2"] = f"Escala:20-69: BAJO  70-79: BASICO  80-89: ALTO  90-100: SUPERIOR"
    ws["A2"].font = _font(bold=True, size=9)
    ws.merge_cells("A2:C2")

    PCT_MAP = [10, 5, 10, 5, None, 10, 10, 30, 10, None, 5, 5, None]
    for col_idx, (_, grupo, skip) in enumerate(COLUMNAS_FIJAS):
        col_num = DATA_START + col_idx
        pct = PCT_MAP[col_idx]
        if skip:
            ws.cell(2, col_num).value = ""
            ws.cell(2, col_num).fill  = _fill(SKIP_COLOR)
        else:
            ws.cell(2, col_num).value = f"{pct}%" if pct else ""
            ws.cell(2, col_num).fill  = _fill(AZUL_CLARO)
            ws.cell(2, col_num).font  = _font(bold=True, size=9)
        ws.cell(2, col_num).alignment = _center()
        ws.cell(2, col_num).border    = brd

    # ── FILA 3: Encabezados columnas ─────────────────────────────
    ws.row_dimensions[3].height = 20
    for col_num, header in [(1,"Código"), (2,"Estudiante"), (3,"No.")]:
        c = ws.cell(3, col_num)
        c.value     = header
        c.font      = _font(bold=True, color=BLANCO, size=10)
        c.fill      = _fill(AZUL_MEDIO)
        c.alignment = _center()
        c.border    = brd

    for col_idx, (abrev, _, skip) in enumerate(COLUMNAS_FIJAS):
        col_num = DATA_START + col_idx
        c = ws.cell(3, col_num)
        c.value     = abrev
        c.font      = _font(bold=True, color=BLANCO if not skip else NEGRO, size=10)
        c.fill      = _fill(SKIP_COLOR if skip else AZUL_MEDIO)
        c.alignment = _center()
        c.border    = brd

    # ── FILAS DE DATOS ────────────────────────────────────────────
    for est_idx, est in enumerate(estudiantes):
        row = 4 + est_idx
        ws.row_dimensions[row].height = 16

        # Código, nombre, número
        ws.cell(row, 1).value     = est.codigo or ""
        ws.cell(row, 1).font      = _font(size=10)
        ws.cell(row, 1).border    = brd
        ws.cell(row, 2).value     = est.nombre
        ws.cell(row, 2).font      = _font(size=10)
        ws.cell(row, 2).border    = brd
        ws.cell(row, 2).alignment = _left()
        ws.cell(row, 3).value     = est_idx + 1
        ws.cell(row, 3).font      = _font(size=10)
        ws.cell(row, 3).border    = brd
        ws.cell(row, 3).alignment = _center()

        # Columnas de notas
        for col_idx, (abrev, _, skip) in enumerate(COLUMNAS_FIJAS):
            col_num = DATA_START + col_idx
            c = ws.cell(row, col_num)
            c.border = brd
            if skip:
                # ProT / ConT / ActT → vacío + fondo gris
                c.value = ""
                c.fill  = _fill(SKIP_COLOR)
            else:
                # Buscar nota en notas_dict por nombre de categoría
                valor = None
                for cat in categorias:
                    if cat_col_map.get(cat.id) == col_idx:
                        valor = notas_dict.get((est.id, cat.id))
                        break
                c.value     = valor if valor is not None else ""
                c.alignment = _center()
                c.font      = _font(size=10)

    # ── Anchos de columna ─────────────────────────────────────────
    ws.column_dimensions["A"].width = 14   # Código
    ws.column_dimensions["B"].width = 38   # Nombre
    ws.column_dimensions["C"].width = 5    # No.
    note_col_widths = [8, 8, 8, 8, 7, 8, 8, 8, 8, 7, 8, 8, 7]
    for i, w in enumerate(note_col_widths):
        ws.column_dimensions[get_column_letter(DATA_START + i)].width = w

    # ── Info del archivo ──────────────────────────────────────────
    wb.properties.title   = f"Notas {materia.nombre} P{periodo}"
    wb.properties.creator = "Gestor Académico"

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output